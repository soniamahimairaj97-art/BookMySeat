from datetime import date, timedelta
from io import BytesIO

from conftest import auth_headers, login, make_employee
from openpyxl import load_workbook

from app.models import DailyBooking, Holiday


def _next_weekday(start_days_ahead=1):
    d = date.today() + timedelta(days=start_days_ahead)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def test_export_rows_are_dates_columns_are_employees_with_row_totals(client, db, team, manager, employee):
    d1 = _next_weekday()
    d2 = d1 + timedelta(days=1)
    while d2.weekday() >= 5:
        d2 += timedelta(days=1)

    db.add(DailyBooking(employee_id=employee.id, booking_date=d1, status="WFO", slot_number=1))
    db.add(DailyBooking(employee_id=employee.id, booking_date=d2, status="WFH"))
    db.commit()

    token = login(client, manager.email)
    resp = client.get(
        "/export",
        params={"start": d1.isoformat(), "end": d2.isoformat()},
        headers=auth_headers(token),
    )
    assert resp.status_code == 200, resp.text

    wb = load_workbook(BytesIO(resp.content))
    ws = wb["Attendance"]

    # Column A holds BookingDate; row 2 holds employee names; data starts row 3.
    assert ws.cell(row=1, column=1).value == "BookingDate"
    header_names = [c.value for c in ws[2]]
    assert employee.name in header_names
    assert manager.name in header_names
    emp_col = header_names.index(employee.name) + 1

    row1_dates = [ws.cell(row=r, column=1).value for r in range(3, 3 + 10)]
    assert d1.isoformat() in row1_dates
    assert d2.isoformat() in row1_dates

    row_for_d1 = next(r for r in range(3, 12) if ws.cell(row=r, column=1).value == d1.isoformat())
    assert ws.cell(row=row_for_d1, column=emp_col).value == "WFO"

    # Row-end per-day totals (WFO/WFH/L/A) exist and count everyone for that day.
    header_row2 = [c.value for c in ws[2]]
    wfo_total_col = header_row2.index("WFO total") + 1
    assert ws.cell(row=row_for_d1, column=wfo_total_col).value == 1

    # Freeze panes below the two header rows, right of the date column.
    assert ws.freeze_panes == "B3"

    # A weekend between d1/d2 (if any) must not appear as its own row.
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.startswith("TOTAL"):
            continue
        if val:
            assert date.fromisoformat(val).weekday() < 5


def test_export_excludes_holiday_rows(client, db, team, manager):
    d1 = _next_weekday()
    holiday_day = d1 + timedelta(days=1)
    while holiday_day.weekday() >= 5:
        holiday_day += timedelta(days=1)
    db.add(Holiday(holiday_date=holiday_day, name="Founder's Day"))
    db.commit()

    token = login(client, manager.email)
    resp = client.get(
        "/export",
        params={"start": d1.isoformat(), "end": holiday_day.isoformat()},
        headers=auth_headers(token),
    )
    assert resp.status_code == 200

    wb = load_workbook(BytesIO(resp.content))
    ws = wb["Attendance"]
    dates_in_sheet = [ws.cell(row=r, column=1).value for r in range(3, ws.max_row + 1)]
    assert holiday_day.isoformat() not in dates_in_sheet
