import calendar
from datetime import date, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.auth import current_user, hash_password, require_role
from app.database import get_db
from app.models import DailyBooking, Employee, Holiday, Team
from app.schemas import EmployeeIn, EmployeeOut, HolidayIn, HolidayOut

router = APIRouter(tags=["admin"])


@router.get("/employees", response_model=list[EmployeeOut])
def list_employees(
    _user: Employee = Depends(current_user),
    db: Session = Depends(get_db),
):
    return db.query(Employee).order_by(Employee.team_id, Employee.name).all()


@router.get("/holidays", response_model=list[HolidayOut])
def list_holidays(
    _user: Employee = Depends(current_user),
    db: Session = Depends(get_db),
):
    return db.query(Holiday).order_by(Holiday.holiday_date).all()


@router.post("/employees", response_model=EmployeeOut)
def add_employee(
    req: EmployeeIn,
    _manager: Employee = Depends(require_role("admin", "manager")),
    db: Session = Depends(get_db),
):
    if not db.query(Team).filter_by(id=req.team_id, is_active=True).first():
        raise HTTPException(400, "Unknown or inactive team.")
    if db.query(Employee).filter_by(email=req.email).first():
        raise HTTPException(409, "An employee with this email already exists.")

    employee = Employee(
        name=req.name,
        email=req.email,
        team_id=req.team_id,
        role=req.role,
        hashed_password=hash_password(req.password),
        is_active=True,
    )
    db.add(employee)
    db.commit()
    db.refresh(employee)
    return employee


@router.post("/holidays", response_model=HolidayOut)
def add_holiday(
    req: HolidayIn,
    manager: Employee = Depends(require_role("admin", "manager")),
    db: Session = Depends(get_db),
):
    if db.query(Holiday).filter_by(holiday_date=req.holiday_date).first():
        raise HTTPException(409, "A holiday is already set for this date.")

    holiday = Holiday(holiday_date=req.holiday_date, name=req.name, created_by=manager.id)
    db.add(holiday)
    db.commit()
    db.refresh(holiday)
    return holiday


@router.delete("/holidays/{holiday_date}")
def delete_holiday(
    holiday_date: date,
    _manager: Employee = Depends(require_role("admin", "manager")),
    db: Session = Depends(get_db),
):
    holiday = db.query(Holiday).filter_by(holiday_date=holiday_date).first()
    if not holiday:
        raise HTTPException(404, "No holiday set for this date.")
    db.delete(holiday)
    db.commit()
    return {"deleted": True}


@router.get("/export")
def export_excel(
    start: date | None = Query(default=None),
    end: date | None = Query(default=None),
    _manager: Employee = Depends(require_role("admin", "manager")),
    db: Session = Depends(get_db),
):
    today = date.today()
    if start is None:
        start = today.replace(day=1)
    if end is None:
        last_day = calendar.monthrange(today.year, today.month)[1]
        end = today.replace(day=last_day)
    if end < start:
        raise HTTPException(400, "end date must not be before start date.")

    holidays = {
        h.holiday_date
        for h in db.query(Holiday).filter(Holiday.holiday_date.between(start, end)).all()
    }
    working_days = [
        start + timedelta(n)
        for n in range((end - start).days + 1)
        if (start + timedelta(n)).weekday() < 5 and (start + timedelta(n)) not in holidays
    ]

    employees = db.query(Employee).filter_by(is_active=True).order_by(Employee.team_id, Employee.name).all()
    bookings = db.query(DailyBooking).filter(
        DailyBooking.booking_date >= start, DailyBooking.booking_date <= end
    ).all()
    booking_map = {(b.employee_id, b.booking_date): b.status for b in bookings}
    teams = {t.id: t.label for t in db.query(Team).all()}

    teams_with_members: dict[str, list] = {}
    for emp in employees:
        teams_with_members.setdefault(teams.get(emp.team_id, ""), []).append(emp)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    total_cols = ["WFO total", "WFH total", "L total", "A total"]
    emp_col_start = 2  # column B — column A is BookingDate

    # Row 1: team header (merged over each team's employee columns) + "Day Totals" header
    ws.cell(row=1, column=1, value="BookingDate")
    col = emp_col_start
    for team_label, members in teams_with_members.items():
        ws.cell(row=1, column=col, value=team_label)
        if len(members) > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(members) - 1)
        col += len(members)
    day_totals_start = col
    ws.cell(row=1, column=day_totals_start, value="Day Totals")
    ws.merge_cells(start_row=1, start_column=day_totals_start, end_row=1, end_column=day_totals_start + 3)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # Row 2: employee names + per-day total column labels
    col = emp_col_start
    for emp in employees:
        ws.cell(row=2, column=col, value=emp.name)
        col += 1
    for offset, label in enumerate(total_cols):
        ws.cell(row=2, column=day_totals_start + offset, value=label)

    # Data rows — one per working day (Y axis), one column per employee (X axis)
    per_employee_totals: dict[int, dict[str, int]] = {emp.id: {"WFO": 0, "WFH": 0, "L": 0, "A": 0} for emp in employees}
    row_num = 3
    for d in working_days:
        row = [d.isoformat()]
        day_counts = {"WFO": 0, "WFH": 0, "L": 0, "A": 0}
        for emp in employees:
            status = booking_map.get((emp.id, d), "A")
            row.append(status)
            day_counts[status] = day_counts.get(status, 0) + 1
            per_employee_totals[emp.id][status] += 1
        row += [day_counts["WFO"], day_counts["WFH"], day_counts["L"], day_counts["A"]]
        ws.append(row)
        row_num += 1

    # Footer block — per-employee totals under each column (replaces old per-employee total columns)
    for status in ("WFO", "WFH", "L", "A"):
        ws.cell(row=row_num, column=1, value=f"TOTAL — {status}")
        col = emp_col_start
        for emp in employees:
            ws.cell(row=row_num, column=col, value=per_employee_totals[emp.id][status])
            col += 1
        row_num += 1

    ws.freeze_panes = "B3"

    totals_ws = wb.create_sheet("Team Totals")
    totals_ws.append(["Team", "WFO", "WFH", "L", "A"])
    for team_label, members in teams_with_members.items():
        counts = {"WFO": 0, "WFH": 0, "L": 0, "A": 0}
        for emp in members:
            for status in counts:
                counts[status] += per_employee_totals[emp.id][status]
        totals_ws.append([team_label, counts["WFO"], counts["WFH"], counts["L"], counts["A"]])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"attendance_{start.isoformat()}_{end.isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
